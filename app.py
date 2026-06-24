from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import os
from datetime import datetime, timedelta
import uuid
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook
import csv
import io
import zipfile

# Shared helpers split out of this file
from db import get_db
from auth_utils import login_required, client_login_required

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-change-this'
app.config['UPLOAD_FOLDER'] = 'static/uploads'
# 16 MB max upload size. Previously written as `16 * 16 * 1024`, which is
# actually 256 KB (16*16*1024 = 262144 bytes) — any photo over a quarter of
# a megabyte, which is nearly every phone photo, was getting rejected by
# Flask before the upload route even ran, surfacing as a raw "413 Request
# Entity Too Large" page rather than anything from this app's own code.
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Ensure upload directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)


@app.errorhandler(413)
def file_too_large(e):
    """Flask's default 413 page is a bare, unstyled error with no context.
    Send the person back to whatever form they were on with a clear
    explanation instead — request.referrer is the page that submitted the
    oversized upload, so this naturally returns to new.html, edit.html, or
    wherever else a file is uploaded from."""
    flash('That file is too large — please choose a photo under 16MB.')
    return redirect(request.referrer or url_for('dashboard'))


def init_template_client_column():
    """Add a nullable client_id column to workout_templates if missing.

    NULL  = universal template (any client).
    value = template that belongs to one specific client.
    Idempotent; safe to run on every startup (also under WSGI).
    """
    conn = get_db()
    try:
        conn.execute('ALTER TABLE workout_templates ADD COLUMN client_id TEXT')
        conn.commit()
        print('[templates] added client_id column to workout_templates')
    except Exception:
        pass
    finally:
        conn.close()


def init_workout_type_column():
    """Add a workout_type column to workout_logs if missing.

    'weightlifting' or 'cardio'. Existing rows (logged before this column
    existed) are backfilled to 'weightlifting' so they keep showing up
    exactly where they did before. A weightlifting and a cardio workout
    can now coexist on the same (client_id, workout_date) without
    overriding or merging with each other — every query that used to key
    on date alone now keys on (date, workout_type) too.
    Idempotent; safe to run on every startup (also under WSGI).
    """
    conn = get_db()
    try:
        conn.execute("ALTER TABLE workout_logs ADD COLUMN workout_type TEXT NOT NULL DEFAULT 'weightlifting'")
        conn.commit()
        print('[workouts] added workout_type column to workout_logs')
    except Exception:
        pass
    finally:
        conn.close()


def init_template_type_column():
    """Add a workout_type column to workout_templates if missing.

    Same idea as workout_logs' workout_type — existing templates are
    weightlifting (sets are weight/reps), backfilled accordingly. Cardio
    templates store distance/duration sets instead.
    Idempotent; safe to run on every startup (also under WSGI).
    """
    conn = get_db()
    try:
        conn.execute("ALTER TABLE workout_templates ADD COLUMN workout_type TEXT NOT NULL DEFAULT 'weightlifting'")
        conn.commit()
        print('[templates] added workout_type column to workout_templates')
    except Exception:
        pass
    finally:
        conn.close()


@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    if 'client_account_id' in session:
        return redirect(url_for('client_portal'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']

        conn = get_db()
        user = conn.execute('SELECT * FROM users WHERE email = ?', (email,)).fetchone()
        conn.close()

        if user and user['password_hash'] == password:
            session['user_id'] = user['id']
            session['user_name'] = user['name']
            # Persisted theme preference (falls back to 'light' for older accounts)
            try:
                session['theme'] = user['theme'] or 'light'
            except (IndexError, KeyError):
                session['theme'] = 'light'
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid email or password')

    return render_template('auth/login.html')


@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        password = request.form['password']
        business_name = request.form.get('business_name', '')

        conn = get_db()
        existing_user = conn.execute('SELECT id FROM users WHERE email = ?', (email,)).fetchone()

        if existing_user:
            flash('Email already registered')
            conn.close()
            return render_template('auth/signup.html')

        user_id = str(uuid.uuid4())
        password_hash = password

        conn.execute('''
            INSERT INTO users (id, name, email, password_hash, business_name, theme, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (user_id, name, email, password_hash, business_name, 'light', datetime.now()))
        conn.commit()
        conn.close()

        flash('Account created successfully! Please log in.')
        return redirect(url_for('login'))

    return render_template('auth/signup.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/api/theme', methods=['POST'])
@login_required
def update_theme():
    data = request.get_json(silent=True) or {}
    theme = data.get('theme')

    if theme not in ('light', 'dark'):
        return jsonify({'error': 'Invalid theme'}), 400

    conn = get_db()
    conn.execute('UPDATE users SET theme = ? WHERE id = ?', (theme, session['user_id']))
    conn.commit()
    conn.close()

    # Keep the session in sync so the next page render is correct
    session['theme'] = theme
    return jsonify({'success': True, 'theme': theme})

@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db()
    trainer_id = session['user_id']

    today = datetime.now().date()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)

    # Get client counts
    total_clients = \
        conn.execute('SELECT COUNT(*) as count FROM clients WHERE trainer_id = ?', (trainer_id,)).fetchone()[
            'count']
    active_clients = conn.execute('SELECT COUNT(*) as count FROM clients WHERE trainer_id = ? AND status = "active"',
                                  (trainer_id,)).fetchone()['count']

    # Sessions scheduled for today (excluding cancelled)
    today_session_count = conn.execute('''
        SELECT COUNT(*) as count FROM sessions
        WHERE trainer_id = ? AND session_date = ? AND status != 'cancelled'
    ''', (trainer_id, today.isoformat())).fetchone()['count']

    # Sessions scheduled for the rest of this week (excluding cancelled)
    week_session_count = conn.execute('''
        SELECT COUNT(*) as count FROM sessions
        WHERE trainer_id = ? AND session_date BETWEEN ? AND ? AND status != 'cancelled'
    ''', (trainer_id, week_start.isoformat(), week_end.isoformat())).fetchone()['count']

    # Get upcoming sessions for the "Upcoming Sessions" card
    recent_sessions = conn.execute('''
        SELECT s.*, c.name as client_name
        FROM sessions s
        JOIN clients c ON s.client_id = c.id
        WHERE s.trainer_id = ? AND s.session_date >= date('now') AND s.status != 'cancelled'
        ORDER BY s.session_date, s.start_time
        LIMIT 5
    ''', (trainer_id,)).fetchall()

    # Recent client-portal activity (workouts, weight, sleep, nutrition logs)
    recent_activity = conn.execute('''
        SELECT id, client_id, client_name, category, action, detail, created_at
        FROM activity_log
        WHERE trainer_id = ?
        ORDER BY created_at DESC
        LIMIT 8
    ''', (trainer_id,)).fetchall()

    conn.close()

    return render_template('dashboard/index.html',
                           total_clients=total_clients,
                           active_clients=active_clients,
                           today_session_count=today_session_count,
                           week_session_count=week_session_count,
                           recent_sessions=recent_sessions,
                           recent_activity=recent_activity)


@app.route('/clients')
@login_required
def clients():
    search = request.args.get('search', '')
    status_filter = request.args.get('status', '')
    sort = request.args.get('sort', '')

    conn = get_db()

    query = '''
        SELECT c.*,
               (SELECT weight FROM weight_logs WHERE client_id = c.id ORDER BY date DESC LIMIT 1) as latest_weight
        FROM clients c
        WHERE c.trainer_id = ?
    '''
    params = [session['user_id']]

    if search:
        query += ' AND (c.name LIKE ? OR c.email LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%'])

    if status_filter:
        query += ' AND c.status = ?'
        params.append(status_filter)

    # Whitelist the sort options — never interpolate raw user input into SQL.
    sort_clauses = {
        'name_asc':  ' ORDER BY c.name COLLATE NOCASE ASC',
        'name_desc': ' ORDER BY c.name COLLATE NOCASE DESC',
    }
    query += sort_clauses.get(sort, ' ORDER BY c.created_at DESC')

    clients = conn.execute(query, params).fetchall()
    conn.close()

    return render_template('dashboard/clients.html', clients=clients,
                           search=search, status_filter=status_filter, sort=sort)


@app.route('/clients/new', methods=['GET', 'POST'])
@login_required
def new_client():
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        phone = request.form.get('phone', '')
        age = request.form.get('age', type=int)
        gender = request.form.get('gender', '')
        weight = None
        height = request.form.get('height', '')
        status = request.form['status']
        notes = request.form.get('notes', '')

        # Handle photo upload
        photo_url = None
        if 'photo' in request.files:
            file = request.files['photo']
            if file and file.filename:
                filename = secure_filename(f"{uuid.uuid4()}_{file.filename}")
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                photo_url = f"uploads/{filename}"

        client_id = str(uuid.uuid4())

        conn = get_db()
        conn.execute('''
            INSERT INTO clients (id, trainer_id, name, email, phone, age, gender, weight, height, status, notes, photo_url, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (client_id, session['user_id'], name, email, phone, age, gender, weight, height, status, notes, photo_url,
              datetime.now()))
        conn.commit()

        generate_portal = request.form.get('generate_portal_code')
        if generate_portal:
            import random, string
            def make_code():
                chars = string.ascii_uppercase + string.digits
                return f"{''.join(random.choices(chars, k=4))}-{''.join(random.choices(chars, k=4))}"

            code = make_code()
            while conn.execute('SELECT id FROM client_accounts WHERE access_code = ?', (code,)).fetchone():
                code = make_code()
            conn.execute('''
                INSERT INTO client_accounts (id, client_id, access_code, is_active, created_at)
                VALUES (?, ?, ?, 1, ?)
            ''', (str(uuid.uuid4()), client_id, code, datetime.now()))
            conn.commit()

        conn.close()

        flash('Client added successfully!')
        return redirect(url_for('clients'))

    return render_template('dashboard/clients/new.html')


@app.route('/clients/<client_id>')
@login_required
def client_detail(client_id):
    app.logger.info(f"[v0] Loading detail page for client_id: {client_id}")
    conn = get_db()
    client = conn.execute('SELECT * FROM clients WHERE id = ? AND trainer_id = ?',
                          (client_id, session['user_id'])).fetchone()

    if not client:
        flash('Client not found')
        return redirect(url_for('clients'))

    upcoming_sessions = conn.execute('''
        SELECT id, session_date, start_time, end_time, session_type, notes, status
        FROM sessions
        WHERE client_id = ? AND session_date >= date('now') AND status != 'completed'
        ORDER BY session_date, start_time
        LIMIT 5
    ''', (client_id,)).fetchall()

    # Get recent workouts
    recent_workouts = conn.execute('''
        SELECT workout_date, COUNT(*) as exercise_count
        FROM workout_logs
        WHERE client_id = ?
        GROUP BY workout_date
        ORDER BY workout_date DESC
        LIMIT 5
    ''', (client_id,)).fetchall()

    weight_history_rows = conn.execute('''
        SELECT id, date, weight, notes
        FROM weight_logs
        WHERE client_id = ?
        ORDER BY date DESC
        LIMIT 10
    ''', (client_id,)).fetchall()

    weight_history = [dict(row) for row in weight_history_rows]

    latest_weight_row = conn.execute('''
        SELECT weight, date
        FROM weight_logs
        WHERE client_id = ?
        ORDER BY date DESC
        LIMIT 1
    ''', (client_id,)).fetchone()

    app.logger.info(f"[v0] Latest weight query for client {client_id}")
    app.logger.info(f"[v0] Latest weight row type: {type(latest_weight_row)}")
    app.logger.info(f"[v0] Latest weight row content: {latest_weight_row}")

    if latest_weight_row:
        latest_weight = dict(latest_weight_row)
        app.logger.info(f"[v0] Latest weight dict keys: {latest_weight.keys()}")
        app.logger.info(
            f"[v0] Latest weight dict values: weight={latest_weight.get('weight')}, date={latest_weight.get('date')}")
    else:
        latest_weight = None
        app.logger.info(f"[v0] No weight logs found for client {client_id}")

    client_notes = conn.execute('''
        SELECT id, note_text, created_at
        FROM client_notes
        WHERE client_id = ?
        ORDER BY created_at DESC
        LIMIT 5
    ''', (client_id,)).fetchall()

    portal_account_row = conn.execute('''
        SELECT access_code, is_active, last_login
        FROM client_accounts WHERE client_id = ?
    ''', (client_id,)).fetchone()
    portal_account = dict(portal_account_row) if portal_account_row else None

    conn.close()

    app.logger.info(f"[v0] Rendering template with latest_weight: {latest_weight}")
    return render_template('dashboard/clients/detail.html',
                           client=client,
                           upcoming_sessions=upcoming_sessions,
                           recent_workouts=recent_workouts,
                           weight_history=weight_history,
                           latest_weight=latest_weight,
                           client_notes=client_notes,
                           portal_account=portal_account)


@app.route('/clients/<client_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_client(client_id):
    conn = get_db()
    client = conn.execute('''
        SELECT * FROM clients WHERE id = ? AND trainer_id = ?
    ''', (client_id, session['user_id'])).fetchone()

    if not client:
        conn.close()
        flash('Client not found')
        return redirect(url_for('clients'))

    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        phone = request.form.get('phone', '')
        age = request.form.get('age', type=int)
        gender = request.form.get('gender', '')
        height = request.form.get('height', '')
        status = request.form['status']
        notes = request.form.get('notes', '')

        # Handle photo upload
        photo_url = client['photo_url']  # Keep existing photo by default
        if 'photo' in request.files:
            file = request.files['photo']
            if file and file.filename:
                # Delete old photo if it exists
                if photo_url and os.path.exists(os.path.join('static', photo_url)):
                    os.remove(os.path.join('static', photo_url))

                filename = secure_filename(f"{uuid.uuid4()}_{file.filename}")
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                photo_url = f"uploads/{filename}"

        conn.execute('''
            UPDATE clients
            SET name = ?, email = ?, phone = ?, age = ?, gender = ?, height = ?,
                status = ?, notes = ?, photo_url = ?
            WHERE id = ? AND trainer_id = ?
        ''', (name, email, phone, age, gender, height, status, notes, photo_url,
              client_id, session['user_id']))
        conn.commit()
        conn.close()

        return redirect(url_for('client_detail', client_id=client_id))

    latest_weight_log = conn.execute('''
        SELECT weight, date FROM weight_logs
        WHERE client_id = ?
        ORDER BY date DESC
        LIMIT 1
    ''', (client_id,)).fetchone()

    latest_weight = latest_weight_log['weight'] if latest_weight_log else None
    latest_weight_date = latest_weight_log['date'] if latest_weight_log else None

    portal_account_row = conn.execute('''
        SELECT access_code, is_active, last_login
        FROM client_accounts WHERE client_id = ?
    ''', (client_id,)).fetchone()
    portal_account = dict(portal_account_row) if portal_account_row else None

    conn.close()
    return render_template('dashboard/clients/edit.html',
                           client=client,
                           latest_weight=latest_weight,
                           latest_weight_date=latest_weight_date,
                           portal_account=portal_account)

@app.route('/clients/<client_id>/delete', methods=['POST'])
@login_required
def delete_client(client_id):
    conn = get_db()

    # Verify client belongs to logged-in trainer
    client = conn.execute('''
        SELECT * FROM clients WHERE id = ? AND trainer_id = ?
    ''', (client_id, session['user_id'])).fetchone()

    if not client:
        conn.close()
        return jsonify({'error': 'Client not found'}), 404

    # Delete client photo if it exists
    if client['photo_url'] and os.path.exists(os.path.join('static', client['photo_url'])):
        try:
            os.remove(os.path.join('static', client['photo_url']))
        except:
            pass

    conn.execute('DELETE FROM weight_logs WHERE client_id = ?', (client_id,))
    conn.execute('DELETE FROM workout_logs WHERE client_id = ?', (client_id,))
    conn.execute('DELETE FROM sessions WHERE client_id = ?', (client_id,))
    conn.execute('DELETE FROM client_notes WHERE client_id = ?', (client_id,))
    conn.execute('DELETE FROM client_accounts WHERE client_id = ?', (client_id,))  # ← add here
    conn.execute('DELETE FROM clients WHERE id = ? AND trainer_id = ?', (client_id, session['user_id']))

    conn.commit()
    conn.close()

    return jsonify({'success': True}), 200


@app.route('/clients/<client_id>/workouts', methods=['GET', 'POST'])
@login_required
def client_workouts(client_id):
    conn = get_db()
    client = conn.execute('SELECT * FROM clients WHERE id = ? AND trainer_id = ?',
                          (client_id, session['user_id'])).fetchone()

    if not client:
        conn.close()
        flash('Client not found')
        return redirect(url_for('clients'))

    if request.method == 'POST':
        workout_date = request.form['date']
        workout_type = request.form.get('workout_type', 'weightlifting')
        if workout_type not in ('weightlifting', 'cardio'):
            workout_type = 'weightlifting'
        exercises = request.form.getlist('exercise_name[]')
        notes_list = request.form.getlist('exercise_notes[]')
        workout_tags = request.form.get('workout_tags', '')
        override = request.form.get('override', 'false') == 'true'

        # Cardio workouts always carry a "Cardio" tag that can't be removed
        # from the form — it's added here server-side regardless of what was
        # submitted, so the only way to get rid of it is deleting the workout.
        if workout_type == 'cardio':
            tag_list = [t.strip() for t in workout_tags.split(',') if t.strip()]
            if 'Cardio' not in tag_list:
                tag_list.append('Cardio')
            workout_tags = ','.join(tag_list)

        if not override:
            existing = conn.execute(
                'SELECT 1 FROM workout_logs WHERE client_id = ? AND workout_date = ? AND workout_type = ? LIMIT 1',
                (client_id, workout_date, workout_type)
            ).fetchone()
            if existing:
                conn.close()
                return jsonify({'conflict': True}), 409

        for i, exercise in enumerate(exercises):
            if exercise.strip():  # Only save non-empty exercises
                if workout_type == 'cardio':
                    distances = request.form.getlist(f'exercise_{i}_distance[]')
                    distance_units = request.form.getlist(f'exercise_{i}_distance_unit[]')
                    durations = request.form.getlist(f'exercise_{i}_duration[]')
                    duration_units = request.form.getlist(f'exercise_{i}_duration_unit[]')
                    set_notes = request.form.getlist(f'exercise_{i}_set_notes[]')

                    exercise_sets = []
                    set_count = max(len(distances), len(durations), 1)
                    for j in range(set_count):
                        distance = distances[j] if j < len(distances) and distances[j] else None
                        duration = durations[j] if j < len(durations) and durations[j] else None
                        exercise_sets.append({
                            'distance': float(distance) if distance else None,
                            'distance_unit': distance_units[j] if j < len(distance_units) and distance_units[j] else None,
                            'duration': float(duration) if duration else None,
                            'duration_unit': duration_units[j] if j < len(duration_units) and duration_units[j] else None,
                            'notes': set_notes[j] if j < len(set_notes) and set_notes[j] else None,
                        })
                    if not exercise_sets:
                        exercise_sets = [{'distance': None, 'distance_unit': None, 'duration': None, 'duration_unit': None, 'notes': None}]

                    total_sets = len(exercise_sets)
                    conn.execute('''
                        INSERT INTO workout_logs (id, client_id, trainer_id, exercise_name, sets, reps, weight, notes, workout_date, sets_data, tags, workout_type, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (str(uuid.uuid4()), client_id, session['user_id'], exercise.strip(),
                          total_sets, None, None,
                          notes_list[i] if i < len(notes_list) else '',
                          workout_date, json.dumps(exercise_sets), workout_tags, workout_type, datetime.now()))
                else:
                    # Get sets for this specific exercise using the new field naming
                    exercise_weights = request.form.getlist(f'exercise_{i}_weight[]')
                    exercise_reps = request.form.getlist(f'exercise_{i}_reps[]')

                    # Build sets data for this exercise
                    exercise_sets = []
                    for j in range(len(exercise_weights)):
                        weight = exercise_weights[j] if j < len(exercise_weights) and exercise_weights[j] else None
                        reps = exercise_reps[j] if j < len(exercise_reps) and exercise_reps[j] else None

                        exercise_sets.append({
                            'weight': float(weight) if weight else None,
                            'reps': int(reps) if reps else None
                        })

                    # If no sets data, create a default set
                    if not exercise_sets:
                        exercise_sets = [{'weight': None, 'reps': None}]

                    # Calculate totals for backward compatibility
                    total_sets = len(exercise_sets)
                    avg_reps = sum(s['reps'] for s in exercise_sets if s['reps']) // len(
                        [s for s in exercise_sets if s['reps']]) if any(s['reps'] for s in exercise_sets) else None
                    avg_weight = sum(s['weight'] for s in exercise_sets if s['weight']) / len(
                        [s for s in exercise_sets if s['weight']]) if any(s['weight'] for s in exercise_sets) else None

                    conn.execute('''
                        INSERT INTO workout_logs (id, client_id, trainer_id, exercise_name, sets, reps, weight, notes, workout_date, sets_data, tags, workout_type, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (str(uuid.uuid4()), client_id, session['user_id'], exercise.strip(),
                          total_sets, avg_reps, avg_weight,
                          notes_list[i] if i < len(notes_list) else '',
                          workout_date, json.dumps(exercise_sets), workout_tags, workout_type, datetime.now()))

        conn.commit()
        conn.close()
        return jsonify({'success': True})

    workouts = conn.execute('''
        SELECT workout_date, workout_type, COUNT(*) as exercise_count
        FROM workout_logs
        WHERE client_id = ?
        GROUP BY workout_date, workout_type
        ORDER BY workout_date DESC, workout_type
    ''', (client_id,)).fetchall()

    conn.close()
    return render_template('dashboard/clients/workouts.html', client=client, workouts=workouts)


@app.route('/clients/<client_id>/workouts/new', methods=['GET', 'POST'])
@login_required
def new_workout(client_id):
    conn = get_db()
    client = conn.execute('SELECT * FROM clients WHERE id = ? AND trainer_id = ?',
                          (client_id, session['user_id'])).fetchone()

    if not client:
        flash('Client not found')
        return redirect(url_for('clients'))

    if request.method == 'POST':
        workout_date = request.form['workout_date']
        exercises = request.form.getlist('exercise_name[]')
        notes_list = request.form.getlist('exercise_notes[]')

        for i, exercise in enumerate(exercises):
            if exercise.strip():  # Only save non-empty exercises
                # Get sets for this specific exercise using the new field naming
                exercise_weights = request.form.getlist(f'exercise_{i}_weight[]')
                exercise_reps = request.form.getlist(f'exercise_{i}_reps[]')

                # Build sets data for this exercise
                exercise_sets = []
                for j in range(len(exercise_weights)):
                    weight = exercise_weights[j] if j < len(exercise_weights) and exercise_weights[j] else None
                    reps = exercise_reps[j] if j < len(exercise_reps) and exercise_reps[j] else None

                    exercise_sets.append({
                        'weight': float(weight) if weight else None,
                        'reps': int(reps) if reps else None
                    })

                # If no sets data, create a default set
                if not exercise_sets:
                    exercise_sets = [{'weight': None, 'reps': None}]

                # Calculate totals for backward compatibility
                total_sets = len(exercise_sets)
                avg_reps = sum(s['reps'] for s in exercise_sets if s['reps']) // len(
                    [s for s in exercise_sets if s['reps']]) if any(s['reps'] for s in exercise_sets) else None
                avg_weight = sum(s['weight'] for s in exercise_sets if s['weight']) / len(
                    [s for s in exercise_sets if s['weight']]) if any(s['weight'] for s in exercise_sets) else None

                conn.execute('''
                    INSERT INTO workout_logs (id, client_id, trainer_id, exercise_name, sets, reps, weight, notes, workout_date, sets_data, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (str(uuid.uuid4()), client_id, session['user_id'], exercise.strip(),
                      total_sets, avg_reps, avg_weight,
                      notes_list[i] if i < len(notes_list) else '',
                      workout_date, json.dumps(exercise_sets), datetime.now()))

        conn.commit()
        conn.close()

        flash('Workout logged successfully!')
        return redirect(url_for('client_workouts', client_id=client_id))

    conn.close()
    return render_template('dashboard/clients/new_workout.html', client=client, today=datetime.now().date())


@app.route('/api/exercises/search')
@login_required
def search_exercises():
    query = request.args.get('q', '').lower()

    conn = get_db()
    exercises = conn.execute('''
        SELECT name, muscle_group, equipment
        FROM exercises
        WHERE LOWER(name) LIKE ?
        ORDER BY name
        LIMIT 10
    ''', (f'%{query}%',)).fetchall()
    conn.close()

    return jsonify([{
        'name': ex['name'],
        'muscle_group': ex['muscle_group'],
        'equipment': ex['equipment']
    } for ex in exercises])


@app.route('/calendar')
@login_required
def calendar():
    # Get current week with offset
    week_offset = request.args.get('week_offset', 0, type=int)

    today = datetime.now().date()
    start_of_week = today - timedelta(days=today.weekday()) + timedelta(weeks=week_offset)
    end_of_week = start_of_week + timedelta(days=6)

    week_dates = []
    for i in range(7):
        week_dates.append(start_of_week + timedelta(days=i))

    user_id = session['user_id']

    conn = get_db()
    sessions_data = conn.execute('''
        SELECT s.*, c.name as client_name
        FROM sessions s
        JOIN clients c ON s.client_id = c.id
        WHERE s.trainer_id = ? AND s.session_date BETWEEN ? AND ?
        ORDER BY s.session_date, s.start_time
    ''', (user_id, start_of_week, end_of_week)).fetchall()

    clients = conn.execute('SELECT id, name FROM clients WHERE trainer_id = ? ORDER BY name', (user_id,)).fetchall()

    conn.close()

    # Organize sessions by day
    week_sessions = {}
    for i in range(7):
        day = start_of_week + timedelta(days=i)
        week_sessions[day.strftime('%Y-%m-%d')] = []

    for session_item in sessions_data:
        day_key = session_item['session_date']
        if day_key in week_sessions:
            week_sessions[day_key].append(session_item)

    return render_template('dashboard/calendar.html',
                           week_sessions=week_sessions,
                           week_dates=week_dates,
                           week_start=start_of_week,
                           week_end=end_of_week,
                           week_offset=week_offset,
                           clients=clients)


@app.route('/activity-stream')
@login_required
def activity_stream():
    """Trainer view: a week of client portal activity plus upcoming sessions."""
    week_offset = request.args.get('week_offset', 0, type=int)

    today = datetime.now().date()
    start_of_week = today - timedelta(days=today.weekday()) + timedelta(weeks=week_offset)
    end_of_week = start_of_week + timedelta(days=6)

    week_dates = [start_of_week + timedelta(days=i) for i in range(7)]

    user_id = session['user_id']
    conn = get_db()

    # Activity events for this trainer within the week.
    # created_at is a full timestamp; compare on its date portion.
    activity_rows = conn.execute('''
        SELECT id, client_id, client_name, category, action, detail, created_at
        FROM activity_log
        WHERE trainer_id = ?
          AND date(created_at) BETWEEN ? AND ?
        ORDER BY created_at DESC
    ''', (user_id, start_of_week.isoformat(), end_of_week.isoformat())).fetchall()

    # Upcoming sessions in the same week (booked on the calendar).
    session_rows = conn.execute('''
        SELECT s.id, s.session_date, s.start_time, s.end_time, s.session_type,
               s.status, c.name AS client_name
        FROM sessions s
        JOIN clients c ON s.client_id = c.id
        WHERE s.trainer_id = ?
          AND s.session_date BETWEEN ? AND ?
        ORDER BY s.session_date, s.start_time
    ''', (user_id, start_of_week.isoformat(), end_of_week.isoformat())).fetchall()

    conn.close()

    # Group both into per-day buckets keyed by YYYY-MM-DD.
    week_activity = {}
    week_sessions = {}
    for d in week_dates:
        key = d.strftime('%Y-%m-%d')
        week_activity[key] = []
        week_sessions[key] = []

    for r in activity_rows:
        # created_at may be 'YYYY-MM-DD HH:MM:SS(.ffffff)'
        day_key = str(r['created_at'])[:10]
        if day_key in week_activity:
            week_activity[day_key].append(dict(r))

    for s in session_rows:
        day_key = s['session_date']
        if day_key in week_sessions:
            week_sessions[day_key].append(dict(s))

    return render_template('dashboard/activity_stream.html',
                           week_activity=week_activity,
                           week_sessions=week_sessions,
                           week_dates=week_dates,
                           week_start=start_of_week,
                           week_end=end_of_week,
                           week_offset=week_offset)


@app.route('/api/sessions', methods=['POST'])
@login_required
def create_session():
    data = request.json
    client_id = data['client_id']
    session_date = data['session_date']
    start_time = data['start_time']
    end_time = data['end_time']
    session_type = data.get('session_type', 'training')
    notes = data.get('notes', '')

    session_id = str(uuid.uuid4())

    user_id = session['user_id']

    conn = get_db()
    conn.execute('''
        INSERT INTO sessions (id, trainer_id, client_id, session_date, start_time, end_time, session_type, status, notes, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, 'scheduled', ?, ?)
    ''', (session_id, user_id, client_id, session_date, start_time, end_time, session_type, notes, datetime.now()))
    conn.commit()
    conn.close()

    return jsonify({'success': True, 'session_id': session_id})


@app.route('/api/sessions/<session_id>', methods=['GET'])
@login_required
def get_session(session_id):
    conn = get_db()
    session_data = conn.execute('''
        SELECT * FROM sessions
        WHERE id = ? AND trainer_id = ?
    ''', (session_id, session['user_id'])).fetchone()

    if not session_data:
        conn.close()
        return jsonify({'error': 'Session not found'}), 404

    conn.close()
    return jsonify(dict(session_data))


@app.route('/api/sessions/<session_id>', methods=['PUT'])
@login_required
def update_session(session_id):
    data = request.json

    conn = get_db()
    # Verify session belongs to current trainer
    session_data = conn.execute('''
        SELECT * FROM sessions
        WHERE id = ? AND trainer_id = ?
    ''', (session_id, session['user_id'])).fetchone()

    if not session_data:
        conn.close()
        return jsonify({'error': 'Session not found'}), 404

    conn.execute('''
        UPDATE sessions
        SET session_date = ?, start_time = ?, end_time = ?, session_type = ?,
            notes = ?, status = ?, updated_at = ?
        WHERE id = ? AND trainer_id = ?
    ''', (data['session_date'], data['start_time'], data['end_time'],
          data['session_type'], data.get('notes', ''),
          data.get('status', 'scheduled'), datetime.now(),
          session_id, session['user_id']))

    conn.commit()
    conn.close()

    return jsonify({'success': True})


@app.route('/api/sessions/<session_id>', methods=['DELETE'])
@login_required
def delete_session(session_id):
    conn = get_db()
    # Verify session belongs to current trainer
    session_data = conn.execute('''
        SELECT * FROM sessions
        WHERE id = ? AND trainer_id = ?
    ''', (session_id, session['user_id'])).fetchone()

    if not session_data:
        conn.close()
        return jsonify({'error': 'Session not found'}), 404

    conn.execute('DELETE FROM sessions WHERE id = ? AND trainer_id = ?',
                 (session_id, session['user_id']))

    conn.commit()
    conn.close()

    return jsonify({'success': True})


@app.route('/api/sessions/<session_id>/complete', methods=['POST'])
@login_required
def complete_session(session_id):
    conn = get_db()
    # Verify session belongs to current trainer
    session_data = conn.execute('''
        SELECT * FROM sessions
        WHERE id = ? AND trainer_id = ?
    ''', (session_id, session['user_id'])).fetchone()

    if not session_data:
        conn.close()
        return jsonify({'error': 'Session not found'}), 404

    conn.execute('''
        UPDATE sessions
        SET status = 'completed', updated_at = ?
        WHERE id = ? AND trainer_id = ?
    ''', (datetime.now(), session_id, session['user_id']))

    conn.commit()
    conn.close()

    return jsonify({'success': True})


@app.route('/api/sessions/<session_id>/cancel', methods=['POST'])
@login_required
def cancel_session(session_id):
    conn = get_db()
    # Verify session belongs to current trainer
    session_data = conn.execute('''
        SELECT * FROM sessions
        WHERE id = ? AND trainer_id = ?
    ''', (session_id, session['user_id'])).fetchone()

    if not session_data:
        conn.close()
        return jsonify({'error': 'Session not found'}), 404

    conn.execute('''
        UPDATE sessions
        SET status = 'cancelled', updated_at = ?
        WHERE id = ? AND trainer_id = ?
    ''', (datetime.now(), session_id, session['user_id']))

    conn.commit()
    conn.close()

    return jsonify({'success': True})


@app.route('/clients/<client_id>/sessions/history')
@login_required
def session_history(client_id):
    conn = get_db()
    client = conn.execute('SELECT * FROM clients WHERE id = ? AND trainer_id = ?',
                          (client_id, session['user_id'])).fetchone()

    if not client:
        flash('Client not found')
        return redirect(url_for('clients'))

    all_sessions = conn.execute('''
        SELECT * FROM sessions
        WHERE client_id = ? AND trainer_id = ?
        ORDER BY session_date DESC, start_time DESC
    ''', (client_id, session['user_id'])).fetchall()

    conn.close()

    return render_template('dashboard/clients/session_history.html',
                           client=client, all_sessions=all_sessions)


@app.route('/clients/<client_id>/workouts/<date>')
@login_required
def workout_detail(client_id, date):
    conn = get_db()
    client = conn.execute('SELECT * FROM clients WHERE id = ? AND trainer_id = ?',
                          (client_id, session['user_id'])).fetchone()

    if not client:
        flash('Client not found')
        return redirect(url_for('clients'))

    # Optional ?type=weightlifting|cardio — when omitted, returns every
    # exercise for that date regardless of type (kept for any older caller
    # that predates workout_type and expects the old, unfiltered behavior).
    workout_type = request.args.get('type')
    if workout_type in ('weightlifting', 'cardio'):
        exercises = conn.execute('''
            SELECT id, exercise_name, sets_data, notes, tags, workout_type
            FROM workout_logs
            WHERE client_id = ? AND workout_date = ? AND trainer_id = ? AND workout_type = ?
            ORDER BY created_at
        ''', (client_id, date, session['user_id'], workout_type)).fetchall()
    else:
        exercises = conn.execute('''
            SELECT id, exercise_name, sets_data, notes, tags, workout_type
            FROM workout_logs
            WHERE client_id = ? AND workout_date = ? AND trainer_id = ?
            ORDER BY created_at
        ''', (client_id, date, session['user_id'])).fetchall()

    conn.close()

    result = []
    for ex in exercises:
        sets_data = None
        if ex['sets_data']:
            try:
                sets_data = json.loads(ex['sets_data'])
            except:
                sets_data = None

        result.append({
            'id': ex['id'],
            'exercise_name': ex['exercise_name'],
            'notes': ex['notes'],
            'sets_data': sets_data,
            'tags': ex['tags'] if ex['tags'] else '',
            'workout_type': ex['workout_type'] if 'workout_type' in ex.keys() else 'weightlifting'
        })

    return jsonify(result)


@app.route('/api/update-workout/<client_id>/<date>', methods=['POST'])
@login_required
def update_workout(client_id, date):
    conn = get_db()
    # Verify client belongs to current trainer
    client = conn.execute('SELECT * FROM clients WHERE id = ? AND trainer_id = ?',
                          (client_id, session['user_id'])).fetchone()

    if not client:
        conn.close()
        return jsonify({'error': 'Client not found'}), 404

    new_date = request.args.get('new_date', date)
    workout_type = request.form.get('workout_type', 'weightlifting')
    if workout_type not in ('weightlifting', 'cardio'):
        workout_type = 'weightlifting'

    exercise_names = request.form.getlist('exercise_name[]')
    notes_list = request.form.getlist('exercise_notes[]')
    workout_tags = request.form.get('workout_tags', '')

    if workout_type == 'cardio':
        tag_list = [t.strip() for t in workout_tags.split(',') if t.strip()]
        if 'Cardio' not in tag_list:
            tag_list.append('Cardio')
        workout_tags = ','.join(tag_list)

    # Delete only this workout's own (date, type) pair — a weightlifting and
    # a cardio workout on the same date are independent and must not affect
    # each other when one of them is edited.
    conn.execute('''
        DELETE FROM workout_logs
        WHERE client_id = ? AND workout_date = ? AND trainer_id = ? AND workout_type = ?
    ''', (client_id, date, session['user_id'], workout_type))

    for i, exercise_name in enumerate(exercise_names):
        if exercise_name.strip():  # Only save non-empty exercises
            if workout_type == 'cardio':
                distances = request.form.getlist(f'exercise_{i}_distance[]')
                distance_units = request.form.getlist(f'exercise_{i}_distance_unit[]')
                durations = request.form.getlist(f'exercise_{i}_duration[]')
                duration_units = request.form.getlist(f'exercise_{i}_duration_unit[]')
                set_notes = request.form.getlist(f'exercise_{i}_set_notes[]')

                exercise_sets = []
                set_count = max(len(distances), len(durations), 1)
                for j in range(set_count):
                    distance = distances[j] if j < len(distances) and distances[j] else None
                    duration = durations[j] if j < len(durations) and durations[j] else None
                    exercise_sets.append({
                        'distance': float(distance) if distance else None,
                        'distance_unit': distance_units[j] if j < len(distance_units) and distance_units[j] else None,
                        'duration': float(duration) if duration else None,
                        'duration_unit': duration_units[j] if j < len(duration_units) and duration_units[j] else None,
                        'notes': set_notes[j] if j < len(set_notes) and set_notes[j] else None,
                    })
                if not exercise_sets:
                    exercise_sets = [{'distance': None, 'distance_unit': None, 'duration': None, 'duration_unit': None, 'notes': None}]

                total_sets = len(exercise_sets)
                notes_val = notes_list[i] if i < len(notes_list) else ''

                conn.execute('''
                    INSERT INTO workout_logs (id, client_id, trainer_id, exercise_name, sets, reps, weight, notes, workout_date, sets_data, tags, workout_type, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (str(uuid.uuid4()), client_id, session['user_id'], exercise_name.strip(),
                      total_sets, None, None, notes_val, new_date, json.dumps(exercise_sets), workout_tags,
                      workout_type, datetime.now()))
            else:
                # Get sets for this specific exercise using the new field naming
                exercise_weights = request.form.getlist(f'exercise_{i}_weight[]')
                exercise_reps = request.form.getlist(f'exercise_{i}_reps[]')

                # Build sets data for this exercise
                exercise_sets = []
                for j in range(len(exercise_weights)):
                    weight = exercise_weights[j] if j < len(exercise_weights) and exercise_weights[j] else None
                    reps = exercise_reps[j] if j < len(exercise_reps) and exercise_reps[j] else None

                    exercise_sets.append({
                        'weight': float(weight) if weight else None,
                        'reps': int(reps) if reps else None
                    })

                # If no sets data, create a default set
                if not exercise_sets:
                    exercise_sets = [{'weight': None, 'reps': None}]

                # Calculate totals for backward compatibility
                total_sets = len(exercise_sets)
                avg_reps = sum(s['reps'] for s in exercise_sets if s['reps']) // len(
                    [s for s in exercise_sets if s['reps']]) if any(s['reps'] for s in exercise_sets) else None
                avg_weight = sum(s['weight'] for s in exercise_sets if s['weight']) / len(
                    [s for s in exercise_sets if s['weight']]) if any(s['weight'] for s in exercise_sets) else None

                notes_val = notes_list[i] if i < len(notes_list) else ''

                conn.execute('''
                    INSERT INTO workout_logs (id, client_id, trainer_id, exercise_name, sets, reps, weight, notes, workout_date, sets_data, tags, workout_type, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (str(uuid.uuid4()), client_id, session['user_id'], exercise_name.strip(),
                      total_sets, avg_reps, avg_weight, notes_val, new_date, json.dumps(exercise_sets), workout_tags,
                      workout_type, datetime.now()))

    conn.commit()
    conn.close()

    return jsonify({'success': True})


@app.route('/api/delete-workout/<client_id>/<date>', methods=['DELETE'])
@login_required
def delete_workout(client_id, date):
    conn = get_db()
    # Verify client belongs to current trainer
    client = conn.execute('SELECT * FROM clients WHERE id = ? AND trainer_id = ?',
                          (client_id, session['user_id'])).fetchone()

    if not client:
        conn.close()
        return jsonify({'error': 'Client not found'}), 404

    # workout_type is required so deleting one workout (e.g. cardio) on a
    # date can never also wipe out a different workout type logged the same
    # day. Falls back to 'weightlifting' only for any pre-existing caller
    # that predates this parameter.
    workout_type = request.args.get('type', 'weightlifting')
    if workout_type not in ('weightlifting', 'cardio'):
        workout_type = 'weightlifting'

    conn.execute('''
        DELETE FROM workout_logs
        WHERE client_id = ? AND workout_date = ? AND trainer_id = ? AND workout_type = ?
    ''', (client_id, date, session['user_id'], workout_type))

    conn.commit()
    conn.close()

    return jsonify({'success': True})


@app.route('/clients/<client_id>/workouts/duplicate', methods=['POST'])
@login_required
def duplicate_workout(client_id):
    app.logger.info(f"[v0] Duplicate workout called for client: {client_id}")

    conn = get_db()
    # Verify client belongs to current trainer
    client = conn.execute('SELECT * FROM clients WHERE id = ? AND trainer_id = ?',
                          (client_id, session['user_id'])).fetchone()

    if not client:
        app.logger.error(f"[v0] Client not found: {client_id}")
        conn.close()
        return jsonify({'error': 'Client not found'}), 404

    data = request.get_json()
    original_date = data.get('original_date')
    new_date = data.get('new_date')
    override = data.get('override', False)
    workout_type = data.get('workout_type', 'weightlifting')
    if workout_type not in ('weightlifting', 'cardio'):
        workout_type = 'weightlifting'

    app.logger.info(f"[v0] Duplicating {workout_type} workout from {original_date} to {new_date}")

    if not original_date or not new_date:
        app.logger.error("[v0] Missing date parameters")
        conn.close()
        return jsonify({'error': 'Missing date parameters'}), 400

    if not override:
        existing = conn.execute(
            'SELECT 1 FROM workout_logs WHERE client_id = ? AND workout_date = ? AND workout_type = ? LIMIT 1',
            (client_id, new_date, workout_type)
        ).fetchone()
        if existing:
            conn.close()
            return jsonify({'conflict': True}), 409

    # Fetch all exercises from the original workout (same type only — a
    # cardio "duplicate" button only ever duplicates the cardio entry for
    # that date, never an unrelated weightlifting entry on the same day).
    exercises = conn.execute('''
        SELECT exercise_name, sets_data, notes, tags
        FROM workout_logs
        WHERE client_id = ? AND workout_date = ? AND trainer_id = ? AND workout_type = ?
        ORDER BY created_at
    ''', (client_id, original_date, session['user_id'], workout_type)).fetchall()

    if not exercises:
        app.logger.warning(f"[v0] No {workout_type} workout found for date {original_date}")
        conn.close()
        return jsonify({'error': 'No workout found for that date'}), 404

    app.logger.info(f"[v0] Found {len(exercises)} exercises to duplicate")

    # Duplicate each exercise to the new date
    default_sets = ([{'distance': None, 'distance_unit': None, 'duration': None, 'duration_unit': None, 'notes': None}]
                     if workout_type == 'cardio' else [{'weight': None, 'reps': None}])
    for exercise in exercises:
        sets_data = exercise['sets_data']
        exercise_sets = json.loads(sets_data) if sets_data else default_sets
        total_sets = len(exercise_sets)

        if workout_type == 'cardio':
            avg_reps = None
            avg_weight = None
        else:
            avg_reps = sum(s['reps'] for s in exercise_sets if s.get('reps')) // len(
                [s for s in exercise_sets if s.get('reps')]) if any(s.get('reps') for s in exercise_sets) else None
            avg_weight = sum(s['weight'] for s in exercise_sets if s.get('weight')) / len(
                [s for s in exercise_sets if s.get('weight')]) if any(s.get('weight') for s in exercise_sets) else None

        conn.execute('''
            INSERT INTO workout_logs (id, client_id, trainer_id, exercise_name, sets, reps, weight, notes, workout_date, sets_data, tags, workout_type, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (str(uuid.uuid4()), client_id, session['user_id'], exercise['exercise_name'],
              total_sets, avg_reps, avg_weight, exercise['notes'], new_date, sets_data, exercise['tags'],
              workout_type, datetime.now()))

    conn.commit()
    conn.close()

    app.logger.info("[v0] Workout duplicated successfully")
    return jsonify({'success': True})


@app.route('/api/weight-log', methods=['POST'])
@login_required
def add_weight_log():
    data = request.json
    client_id = data['client_id']
    date = data['date']
    weight = data['weight']
    notes = data.get('notes', '')
    override = data.get('override', False)

    # Verify client belongs to current trainer
    conn = get_db()
    client = conn.execute('SELECT * FROM clients WHERE id = ? AND trainer_id = ?',
                          (client_id, session['user_id'])).fetchone()

    if not client:
        conn.close()
        return jsonify({'error': 'Client not found'}), 404

    existing = conn.execute(
        'SELECT id, weight, notes FROM weight_logs WHERE client_id = ? AND date = ?',
        (client_id, date)
    ).fetchone()

    if existing and not override:
        conn.close()
        return jsonify({
            'conflict': True,
            'existing': {'id': existing['id'], 'weight': existing['weight'], 'notes': existing['notes']}
        }), 409

    try:
        if existing and override:
            conn.execute(
                'UPDATE weight_logs SET weight = ?, notes = ?, updated_at = ? WHERE id = ? AND client_id = ?',
                (weight, notes, datetime.now(), existing['id'], client_id)
            )
            weight_id = existing['id']
        else:
            weight_id = str(uuid.uuid4())
            conn.execute('''
                INSERT INTO weight_logs (id, client_id, date, weight, notes, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (weight_id, client_id, date, weight, notes, datetime.now(), datetime.now()))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'id': weight_id})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/weight-log/<weight_id>', methods=['PUT'])
@login_required
def update_weight_log(weight_id):
    data = request.json
    date = data['date']
    weight = data['weight']
    notes = data.get('notes', '')

    conn = get_db()

    # Verify weight log belongs to current trainer's client
    weight_log = conn.execute('''
        SELECT wl.* FROM weight_logs wl
        JOIN clients c ON wl.client_id = c.id
        WHERE wl.id = ? AND c.trainer_id = ?
    ''', (weight_id, session['user_id'])).fetchone()

    if not weight_log:
        conn.close()
        return jsonify({'error': 'Weight log not found'}), 404

    try:
        conn.execute('''
            UPDATE weight_logs
            SET date = ?, weight = ?, notes = ?, updated_at = ?
            WHERE id = ?
        ''', (date, weight, notes, datetime.now(), weight_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/weight-log/<weight_id>', methods=['DELETE'])
@login_required
def delete_weight_log(weight_id):
    conn = get_db()

    # Verify weight log belongs to current trainer's client
    weight_log = conn.execute('''
        SELECT wl.* FROM weight_logs wl
        JOIN clients c ON wl.client_id = c.id
        WHERE wl.id = ? AND c.trainer_id = ?
    ''', (weight_id, session['user_id'])).fetchone()

    if not weight_log:
        conn.close()
        return jsonify({'error': 'Weight log not found'}), 404

    try:
        conn.execute('DELETE FROM weight_logs WHERE id = ?', (weight_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/client-notes', methods=['POST'])
@login_required
def add_client_note():
    data = request.json
    client_id = data['client_id']
    note_text = data['note_text']

    # Verify client belongs to current trainer
    conn = get_db()
    client = conn.execute('SELECT * FROM clients WHERE id = ? AND trainer_id = ?',
                          (client_id, session['user_id'])).fetchone()

    if not client:
        conn.close()
        return jsonify({'error': 'Client not found'}), 404

    note_id = str(uuid.uuid4())

    try:
        conn.execute('''
            INSERT INTO client_notes (id, client_id, trainer_id, note_text, created_at)
            VALUES (?, ?, ?, ?, ?)
        ''', (note_id, client_id, session['user_id'], note_text, datetime.now()))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/client-notes/<note_id>', methods=['PUT', 'DELETE'])
@login_required
def manage_client_note(note_id):
    conn = get_db()

    if request.method == 'PUT':
        data = request.get_json()
        note_text = data.get('note_text', '')

        # Verify note belongs to current trainer
        note = conn.execute('''
            SELECT cn.*, c.trainer_id FROM client_notes cn
            JOIN clients c ON cn.client_id = c.id
            WHERE cn.id = ?
        ''', (note_id,)).fetchone()

        if not note or note['trainer_id'] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Note not found'}), 404

        conn.execute('''
            UPDATE client_notes
            SET note_text = ?
            WHERE id = ?
        ''', (note_text, note_id))
        conn.commit()
        conn.close()

        return jsonify({'success': True})

    elif request.method == 'DELETE':
        # Verify note belongs to current trainer
        note = conn.execute('''
            SELECT cn.*, c.trainer_id FROM client_notes cn
            JOIN clients c ON cn.client_id = c.id
            WHERE cn.id = ?
        ''', (note_id,)).fetchone()

        if not note or note['trainer_id'] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Note not found'}), 404

        conn.execute('DELETE FROM client_notes WHERE id = ?', (note_id,))
        conn.commit()
        conn.close()

        return jsonify({'success': True})


@app.route('/api/upload-client-photo', methods=['POST'])
@login_required
def upload_client_photo():
    client_id = request.form['client_id']

    # Verify client belongs to current trainer
    conn = get_db()
    client = conn.execute('SELECT * FROM clients WHERE id = ? AND trainer_id = ?',
                          (client_id, session['user_id'])).fetchone()

    if not client:
        conn.close()
        return jsonify({'error': 'Client not found'}), 404

    if 'photo' not in request.files:
        conn.close()
        return jsonify({'error': 'No photo provided'}), 400

    file = request.files['photo']
    if file and file.filename:
        # Delete old photo if it exists
        if client['photo_url'] and os.path.exists(os.path.join('static', client['photo_url'])):
            os.remove(os.path.join('static', client['photo_url']))

        filename = secure_filename(f"{uuid.uuid4()}_{file.filename}")
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        photo_url = f"uploads/{filename}"

        conn.execute('''
            UPDATE clients SET photo_url = ? WHERE id = ? AND trainer_id = ?
        ''', (photo_url, client_id, session['user_id']))
        conn.commit()
        conn.close()
        return jsonify({'success': True})

    conn.close()
    return jsonify({'error': 'Invalid file'}), 400


@app.route('/clients/<client_id>/weight-logs')
@login_required
def client_weight_logs(client_id):
    conn = get_db()
    client = conn.execute('SELECT * FROM clients WHERE id = ? AND trainer_id = ?',
                          (client_id, session['user_id'])).fetchone()

    if not client:
        flash('Client not found')
        return redirect(url_for('clients'))

    weight_history_rows = conn.execute('''
        SELECT id, date, weight, notes
        FROM weight_logs
        WHERE client_id = ?
        ORDER BY date DESC
    ''', (client_id,)).fetchall()

    weight_history = [dict(row) for row in weight_history_rows]

    conn.close()

    return render_template('dashboard/clients/weight_logs.html',
                           client=client,
                           weight_history=weight_history)


@app.route('/templates')
@login_required
def workout_templates():
    conn = get_db()

    # Universal templates (client_id IS NULL)
    universal_templates = conn.execute('''
        SELECT id, name, created_at, workout_type,
               (SELECT COUNT(*) FROM template_exercises WHERE template_id = workout_templates.id) as exercise_count
        FROM workout_templates
        WHERE trainer_id = ? AND client_id IS NULL
        ORDER BY created_at DESC
    ''', (session['user_id'],)).fetchall()

    # Client cards, each with how many client-specific templates they have
    clients = conn.execute('''
        SELECT c.id, c.name, c.status, c.photo_url,
               (SELECT COUNT(*) FROM workout_templates wt
                WHERE wt.client_id = c.id AND wt.trainer_id = ?) as template_count
        FROM clients c
        WHERE c.trainer_id = ?
        ORDER BY c.name COLLATE NOCASE
    ''', (session['user_id'], session['user_id'])).fetchall()

    conn.close()

    return render_template('dashboard/workout_templates.html',
                           universal_templates=universal_templates,
                           clients=clients)


@app.route('/templates/client/<client_id>')
@login_required
def specific_workout_templates(client_id):
    conn = get_db()

    client = conn.execute('SELECT * FROM clients WHERE id = ? AND trainer_id = ?',
                          (client_id, session['user_id'])).fetchone()
    if not client:
        conn.close()
        flash('Client not found')
        return redirect(url_for('workout_templates'))

    templates = conn.execute('''
        SELECT id, name, created_at,
               (SELECT COUNT(*) FROM template_exercises WHERE template_id = workout_templates.id) as exercise_count
        FROM workout_templates
        WHERE trainer_id = ? AND client_id = ?
        ORDER BY created_at DESC
    ''', (session['user_id'], client_id)).fetchall()

    conn.close()

    return render_template('dashboard/specific_workout_templates.html',
                           client=client, templates=templates)


@app.route('/exports')
@login_required
def exports():
    conn = get_db()
    clients = conn.execute('''
        SELECT id, name
        FROM clients
        WHERE trainer_id = ?
        ORDER BY name
    ''', (session['user_id'],)).fetchall()
    conn.close()

    return render_template('dashboard/exports.html', clients=clients)


def build_client_export_workbook(conn, client_id, trainer_id, export_workouts, export_weight_logs,
                                  export_nutrition_logs, export_sleep_logs):
    """Build one client's export workbook in memory.

    Returns (client_name, excel_bytes) or None if the client doesn't exist
    or doesn't belong to this trainer.
    """
    client = conn.execute('''
        SELECT name FROM clients
        WHERE id = ? AND trainer_id = ?
    ''', (client_id, trainer_id)).fetchone()

    if not client:
        return None

    # Create Excel workbook for this client
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Export workout history
    if export_workouts:
        ws_workouts = wb.create_sheet('Workout History')

        # Get all workouts for this client, ordered by date (oldest first)
        workouts = conn.execute('''
            SELECT workout_date, exercise_name, sets_data, notes, tags, workout_type
            FROM workout_logs
            WHERE client_id = ?
            ORDER BY workout_date ASC, created_at ASC
        ''', (client_id,)).fetchall()

        # Helper function to populate a workout sheet
        def populate_workout_sheet(ws, filtered_workouts, show_tags=True):  # Added show_tags parameter
            current_row = 1
            current_date = None

            for workout in filtered_workouts:
                is_cardio = workout['workout_type'] == 'cardio' if 'workout_type' in workout.keys() else False

                # Add blank row between different workout dates
                if current_date and current_date != workout['workout_date']:
                    current_row += 1

                # Add workout date header
                if current_date != workout['workout_date']:
                    ws.cell(row=current_row, column=1, value=workout['workout_date'])
                    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
                    current_row += 1
                    current_date = workout['workout_date']

                # Add exercise name
                ws.cell(row=current_row, column=1, value=workout['exercise_name'])
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                current_row += 1

                sets_data = json.loads(workout['sets_data']) if workout['sets_data'] else []

                if is_cardio:
                    # Add sets header (cardio: distance / duration / notes)
                    ws.cell(row=current_row, column=1, value='Set')
                    ws.cell(row=current_row, column=2, value='Distance')
                    ws.cell(row=current_row, column=3, value='Duration')
                    ws.cell(row=current_row, column=4, value='Set Notes')
                    for col in range(1, 5):
                        ws.cell(row=current_row, column=col).font = Font(bold=True)
                    current_row += 1

                    for set_num, set_info in enumerate(sets_data, 1):
                        distance = set_info.get('distance')
                        distance_unit = set_info.get('distance_unit') or ''
                        duration = set_info.get('duration')
                        duration_unit = set_info.get('duration_unit') or ''
                        ws.cell(row=current_row, column=1, value=f'Set {set_num}')
                        ws.cell(row=current_row, column=2,
                                value=f'{distance} {distance_unit}'.strip() if distance is not None else '')
                        ws.cell(row=current_row, column=3,
                                value=f'{duration} {duration_unit}'.strip() if duration is not None else '')
                        ws.cell(row=current_row, column=4, value=set_info.get('notes') or '')
                        current_row += 1
                else:
                    # Add sets header (weightlifting: weight / reps)
                    ws.cell(row=current_row, column=1, value='Set')
                    ws.cell(row=current_row, column=2, value='Weight (lbs)')
                    ws.cell(row=current_row, column=3, value='Reps')
                    for col in range(1, 4):
                        ws.cell(row=current_row, column=col).font = Font(bold=True)
                    current_row += 1

                    for set_num, set_info in enumerate(sets_data, 1):
                        ws.cell(row=current_row, column=1, value=f'Set {set_num}')
                        ws.cell(row=current_row, column=2, value=set_info.get('weight', ''))
                        ws.cell(row=current_row, column=3, value=set_info.get('reps', ''))
                        current_row += 1

                # Add notes if present
                if workout['notes']:
                    ws.cell(row=current_row, column=1, value=f"Notes: {workout['notes']}")
                    ws.cell(row=current_row, column=1).font = Font(italic=True)
                    current_row += 1

                if show_tags and workout['tags']:
                    ws.cell(row=current_row, column=1, value=f"Tags: {workout['tags']}")
                    current_row += 1

            # Adjust column widths (wide enough for either layout: weight/reps or distance/duration/notes)
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 30

        populate_workout_sheet(ws_workouts, workouts)

        muscle_groups = ['Chest', 'Back', 'Biceps', 'Triceps', 'Shoulders', 'Legs', 'Core']

        for muscle_group in muscle_groups:
            # Filter workouts that have this muscle group tag
            # Group by workout_date to get unique workout dates first
            workout_dates_with_tag = {}
            for workout in workouts:
                if workout['tags'] and muscle_group in workout['tags']:
                    date = workout['workout_date']
                    if date not in workout_dates_with_tag:
                        workout_dates_with_tag[date] = []
                    workout_dates_with_tag[date].append(workout)

            # If there are workouts with this tag, create a sheet
            if workout_dates_with_tag:
                ws_muscle = wb.create_sheet(muscle_group)

                # Flatten the workouts back into a list for population
                filtered_workouts = []
                for date in sorted(workout_dates_with_tag.keys()):
                    filtered_workouts.extend(workout_dates_with_tag[date])

                populate_workout_sheet(ws_muscle, filtered_workouts, show_tags=False)

        # Cardio gets its own sheet too, same as the muscle groups above —
        # filtered by workout_type (a real column) rather than tag text,
        # since every cardio workout is guaranteed to have that type set
        # even if its tags ever changed.
        cardio_dates = {}
        for workout in workouts:
            is_cardio = workout['workout_type'] == 'cardio' if 'workout_type' in workout.keys() else False
            if is_cardio:
                date = workout['workout_date']
                cardio_dates.setdefault(date, []).append(workout)

        if cardio_dates:
            ws_cardio = wb.create_sheet('Cardio')
            filtered_cardio = []
            for date in sorted(cardio_dates.keys()):
                filtered_cardio.extend(cardio_dates[date])
            populate_workout_sheet(ws_cardio, filtered_cardio, show_tags=False)

    # Export weight logs
    if export_weight_logs:
        ws_weight = wb.create_sheet('Weight Logs')

        # Add headers
        ws_weight.cell(row=1, column=1, value='Date')
        ws_weight.cell(row=1, column=2, value='Weight (lbs)')
        ws_weight.cell(row=1, column=1).font = Font(bold=True)
        ws_weight.cell(row=1, column=2).font = Font(bold=True)

        # Get weight logs ordered by date (oldest first)
        weight_logs = conn.execute('''
            SELECT date, weight
            FROM weight_logs
            WHERE client_id = ?
            ORDER BY date ASC
        ''', (client_id,)).fetchall()

        for idx, log in enumerate(weight_logs, start=2):
            ws_weight.cell(row=idx, column=1, value=log['date'])
            ws_weight.cell(row=idx, column=2, value=log['weight'])

        # Adjust column widths
        ws_weight.column_dimensions['A'].width = 15
        ws_weight.column_dimensions['B'].width = 15

    if export_nutrition_logs:
        ws_nutrition = wb.create_sheet('Nutrition')

        # Add headers
        ws_nutrition.cell(row=1, column=1, value='Date')
        ws_nutrition.cell(row=1, column=2, value='Diet')  # Added Diet column
        ws_nutrition.cell(row=1, column=3, value='Calories')
        ws_nutrition.cell(row=1, column=4, value='Sodium')
        ws_nutrition.cell(row=1, column=5, value='Sat Fat')  # Renamed column
        ws_nutrition.cell(row=1, column=6, value='Notes')  # Added Notes column
        for col in range(1, 7):
            ws_nutrition.cell(row=1, column=col).font = Font(bold=True)

        # Get nutrition logs ordered by date (oldest first)
        nutrition_logs = conn.execute('''
            SELECT date, diet, estimated_calories, estimated_sodium, estimated_saturated_fat, notes
            FROM nutrition_logs
            WHERE client_id = ?
            ORDER BY date ASC
        ''', (client_id,)).fetchall()

        for idx, log in enumerate(nutrition_logs, start=2):
            ws_nutrition.cell(row=idx, column=1, value=log['date'])
            ws_nutrition.cell(row=idx, column=2, value=log['diet'])
            ws_nutrition.cell(row=idx, column=3,
                              value=log['estimated_calories'] if log['estimated_calories'] else '')
            ws_nutrition.cell(row=idx, column=4,
                              value=log['estimated_sodium'] if log['estimated_sodium'] else '')
            ws_nutrition.cell(row=idx, column=5,
                              value=log['estimated_saturated_fat'] if log['estimated_saturated_fat'] else '')
            ws_nutrition.cell(row=idx, column=6, value=log['notes'])

        # Adjust column widths
        ws_nutrition.column_dimensions['A'].width = 15
        ws_nutrition.column_dimensions['B'].width = 30  # Adjusted width for Diet
        ws_nutrition.column_dimensions['C'].width = 15
        ws_nutrition.column_dimensions['D'].width = 15
        ws_nutrition.column_dimensions['E'].width = 15
        ws_nutrition.column_dimensions['F'].width = 40  # Adjusted width for Notes

    if export_sleep_logs:
        ws_sleep = wb.create_sheet('Sleep Logs')

        # Add headers
        ws_sleep.cell(row=1, column=1, value='Date')
        ws_sleep.cell(row=1, column=2, value='Hours')
        ws_sleep.cell(row=1, column=3, value='Notes')
        ws_sleep.cell(row=1, column=1).font = Font(bold=True)
        ws_sleep.cell(row=1, column=2).font = Font(bold=True)
        ws_sleep.cell(row=1, column=3).font = Font(bold=True)

        # Get sleep logs ordered by date (oldest first)
        sleep_logs = conn.execute('''
            SELECT date, hours, notes
            FROM sleep_logs
            WHERE client_id = ?
            ORDER BY date ASC
        ''', (client_id,)).fetchall()

        for idx, log in enumerate(sleep_logs, start=2):
            ws_sleep.cell(row=idx, column=1, value=log['date'])
            ws_sleep.cell(row=idx, column=2, value=log['hours'])
            ws_sleep.cell(row=idx, column=3, value=log['notes'] if log['notes'] else '')

        # Adjust column widths
        ws_sleep.column_dimensions['A'].width = 15
        ws_sleep.column_dimensions['B'].width = 15
        ws_sleep.column_dimensions['C'].width = 40

    # Save workbook to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # Clean filename
    safe_name = "".join(c for c in client['name'] if c.isalnum() or c in (' ', '-', '_')).strip()
    return (safe_name, excel_buffer.read())


@app.route('/exports/generate', methods=['POST'])
@login_required
def generate_export():
    client_ids = request.form.getlist('client_ids')
    export_workouts = 'export_workouts' in request.form
    export_weight_logs = 'export_weight_logs' in request.form
    export_nutrition_logs = 'export_nutrition_logs' in request.form
    export_sleep_logs = 'export_sleep_logs' in request.form

    if not client_ids:
        flash('Please select at least one client')
        return redirect(url_for('exports'))

    if not export_workouts and not export_weight_logs and not export_nutrition_logs and not export_sleep_logs:
        flash('Please select at least one export option')
        return redirect(url_for('exports'))

    conn = get_db()

    # Build each client's workbook; client_ids that don't belong to this
    # trainer (or no longer exist) are silently skipped, same as before.
    built = []
    for client_id in client_ids:
        result = build_client_export_workbook(
            conn, client_id, session['user_id'],
            export_workouts, export_weight_logs, export_nutrition_logs, export_sleep_logs
        )
        if result:
            built.append(result)

    conn.close()

    if not built:
        flash('No matching clients found to export')
        return redirect(url_for('exports'))

    # A single client never needs a ZIP wrapper — just send the .xlsx directly.
    if len(built) == 1:
        safe_name, excel_bytes = built[0]
        return send_file(
            io.BytesIO(excel_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{safe_name}.xlsx'
        )

    # Multiple clients: zip all their workbooks together.
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for safe_name, excel_bytes in built:
            zip_file.writestr(f'{safe_name}.xlsx', excel_bytes)
    zip_buffer.seek(0)

    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'client_exports_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
    )


@app.route('/exports/all-clients', methods=['POST'])
@login_required
def export_all_clients():
    conn = get_db()

    clients = conn.execute('''
        SELECT c.name, c.email, c.phone, c.age, c.gender, c.status, c.id, c.height
        FROM clients c
        WHERE c.trainer_id = ?
        ORDER BY c.name ASC
    ''', (session['user_id'],)).fetchall()

    # Get latest weight log for each client
    clients_with_weight = []
    for client in clients:
        latest_weight = conn.execute('''
            SELECT weight
            FROM weight_logs
            WHERE client_id = ?
            ORDER BY date DESC
            LIMIT 1
        ''', (client['id'],)).fetchone()

        clients_with_weight.append({
            'name': client['name'],
            'email': client['email'],
            'phone': client['phone'],
            'age': client['age'],
            'gender': client['gender'],
            'weight': latest_weight['weight'] if latest_weight else None,
            'height': client['height'],
            'status': client['status']
        })

    conn.close()

    if not clients_with_weight:
        flash('No clients found to export')
        return redirect(url_for('exports'))

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'All Clients'

    # Add headers
    headers = ['Name', 'Email', 'Phone', 'Age', 'Gender', 'Weight (lbs)', 'Height', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row_num, client in enumerate(clients_with_weight, 2):
        ws.cell(row=row_num, column=1, value=client['name'])
        ws.cell(row=row_num, column=2, value=client['email'])
        ws.cell(row=row_num, column=3, value=client['phone'])
        ws.cell(row=row_num, column=4, value=client['age'])
        ws.cell(row=row_num, column=5, value=client['gender'])
        ws.cell(row=row_num, column=6, value=client['weight'])
        ws.cell(row=row_num, column=7, value=client['height'])
        ws.cell(row=row_num, column=8, value=client['status'])

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12

    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return send_file(
        excel_buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'all_clients_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/api/templates', methods=['POST'])
@login_required
def create_template():
    data = request.json
    template_name = data['name']
    exercises = data['exercises']
    client_id = data.get('client_id') or None
    workout_type = data.get('workout_type', 'weightlifting')
    if workout_type not in ('weightlifting', 'cardio'):
        workout_type = 'weightlifting'

    template_id = str(uuid.uuid4())

    conn = get_db()
    try:
        # If assigned to a client, verify the client belongs to this trainer
        if client_id:
            owns = conn.execute('SELECT 1 FROM clients WHERE id = ? AND trainer_id = ?',
                                (client_id, session['user_id'])).fetchone()
            if not owns:
                conn.close()
                return jsonify({'error': 'Client not found'}), 404

        # Create template
        conn.execute('''
            INSERT INTO workout_templates (id, trainer_id, client_id, name, workout_type, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (template_id, session['user_id'], client_id, template_name, workout_type, datetime.now()))

        # Add exercises to template
        for exercise in exercises:
            exercise_id = str(uuid.uuid4())
            conn.execute('''
                INSERT INTO template_exercises (id, template_id, exercise_name, sets_data, notes, exercise_order)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (exercise_id, template_id, exercise['name'],
                  json.dumps(exercise['sets']), exercise.get('notes', ''), exercise['order']))

        conn.commit()
        conn.close()
        return jsonify({'success': True, 'template_id': template_id})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/templates/<template_id>', methods=['GET'])
@login_required
def get_template(template_id):
    conn = get_db()
    template = conn.execute('''
        SELECT * FROM workout_templates
        WHERE id = ? AND trainer_id = ?
    ''', (template_id, session['user_id'])).fetchone()

    if not template:
        conn.close()
        return jsonify({'error': 'Template not found'}), 404

    exercises = conn.execute('''
        SELECT exercise_name, sets_data, notes, exercise_order
        FROM template_exercises
        WHERE template_id = ?
        ORDER BY exercise_order
    ''', (template_id,)).fetchall()

    conn.close()

    return jsonify({
        'id': template['id'],
        'name': template['name'],
        'client_id': template['client_id'] if 'client_id' in template.keys() else None,
        'workout_type': template['workout_type'] if 'workout_type' in template.keys() else 'weightlifting',
        'exercises': [{
            'name': ex['exercise_name'],
            'sets': json.loads(ex['sets_data']) if ex['sets_data'] else [],
            'notes': ex['notes'],
            'order': ex['exercise_order']
        } for ex in exercises]
    })


@app.route('/api/templates/<template_id>', methods=['DELETE'])
@login_required
def delete_template(template_id):
    conn = get_db()
    template = conn.execute('''
        SELECT * FROM workout_templates
        WHERE id = ? AND trainer_id = ?
    ''', (template_id, session['user_id'])).fetchone()

    if not template:
        conn.close()
        return jsonify({'error': 'Template not found'}), 404

    try:
        conn.execute('DELETE FROM template_exercises WHERE template_id = ?', (template_id,))
        conn.execute('DELETE FROM workout_templates WHERE id = ? AND trainer_id = ?',
                     (template_id, session['user_id']))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/templates/<template_id>', methods=['PUT'])
@login_required
def update_template(template_id):
    conn = get_db()
    template = conn.execute('''
        SELECT * FROM workout_templates
        WHERE id = ? AND trainer_id = ?
    ''', (template_id, session['user_id'])).fetchone()

    if not template:
        conn.close()
        return jsonify({'error': 'Template not found'}), 404

    data = request.json
    template_name = data['name']
    exercises = data['exercises']
    client_id = data.get('client_id') or None
    # A template's workout_type is fixed at creation — the UPDATE below never
    # touches that column, so editing only ever changes name/client/exercises,
    # never converts a weightlifting template into a cardio one (or vice
    # versa), since the exercise shapes aren't compatible.

    try:
        if client_id:
            owns = conn.execute('SELECT 1 FROM clients WHERE id = ? AND trainer_id = ?',
                                (client_id, session['user_id'])).fetchone()
            if not owns:
                conn.close()
                return jsonify({'error': 'Client not found'}), 404

        # Update template name + client assignment
        conn.execute('''
            UPDATE workout_templates
            SET name = ?, client_id = ?, updated_at = ?
            WHERE id = ? AND trainer_id = ?
        ''', (template_name, client_id, datetime.now(), template_id, session['user_id']))

        # Delete existing exercises
        conn.execute('DELETE FROM template_exercises WHERE template_id = ?', (template_id,))

        # Add updated exercises
        for exercise in exercises:
            exercise_id = str(uuid.uuid4())
            conn.execute('''
                INSERT INTO template_exercises (id, template_id, exercise_name, sets_data, notes, exercise_order)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (exercise_id, template_id, exercise['name'],
                  json.dumps(exercise['sets']), exercise.get('notes', ''), exercise['order']))

        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/templates/list')
@login_required
def list_templates():
    """For the workout-logging import dropdown.

    ?client_id=<id> → universal templates PLUS that client's own templates.
    (no param)       → universal templates only.
    ?type=weightlifting|cardio → only templates of that workout type
    (the import picker passes this so a cardio template never gets offered
    inside the weightlifting import flow, or vice versa).
    """
    client_id = request.args.get('client_id')
    workout_type = request.args.get('type')
    if workout_type not in ('weightlifting', 'cardio'):
        workout_type = None
    conn = get_db()

    if client_id:
        query = '''
            SELECT id, name, client_id, workout_type
            FROM workout_templates
            WHERE trainer_id = ? AND (client_id IS NULL OR client_id = ?)
        '''
        params = [session['user_id'], client_id]
    else:
        query = '''
            SELECT id, name, client_id, workout_type
            FROM workout_templates
            WHERE trainer_id = ? AND client_id IS NULL
        '''
        params = [session['user_id']]

    if workout_type:
        query += ' AND workout_type = ?'
        params.append(workout_type)

    query += ' ORDER BY client_id IS NULL, name COLLATE NOCASE' if client_id else ' ORDER BY name COLLATE NOCASE'

    templates = conn.execute(query, params).fetchall()
    conn.close()

    return jsonify([{
        'id': t['id'],
        'name': t['name'],
        'is_client_specific': bool(t['client_id']),
        'workout_type': t['workout_type'] if 'workout_type' in t.keys() else 'weightlifting'
    } for t in templates])


@app.route('/clients/<client_id>/nutrition-logs')
@login_required
def client_nutrition_logs(client_id):
    conn = get_db()
    client = conn.execute('SELECT * FROM clients WHERE id = ? AND trainer_id = ?',
                          (client_id, session['user_id'])).fetchone()

    if not client:
        flash('Client not found')
        return redirect(url_for('clients'))

    nutrition_history_rows = conn.execute('''
        SELECT id, date, diet, estimated_calories, estimated_sodium, estimated_saturated_fat, notes
        FROM nutrition_logs
        WHERE client_id = ?
        ORDER BY date DESC
    ''', (client_id,)).fetchall()

    nutrition_history = [dict(row) for row in nutrition_history_rows]

    conn.close()

    return render_template('dashboard/clients/nutrition_logs.html',
                           client=client,
                           nutrition_history=nutrition_history)


@app.route('/api/nutrition-log', methods=['POST'])
@login_required
def add_nutrition_log():
    data = request.json
    client_id = data['client_id']
    date = data['date']
    diet = data['diet']
    estimated_calories = data.get('estimated_calories')
    estimated_protein = data.get('estimated_protein')
    estimated_sodium = data.get('estimated_sodium')
    estimated_saturated_fat = data.get('estimated_saturated_fat')
    notes = data.get('notes', '')
    override = data.get('override', False)

    # Verify client belongs to current trainer
    conn = get_db()
    client = conn.execute('SELECT * FROM clients WHERE id = ? AND trainer_id = ?',
                          (client_id, session['user_id'])).fetchone()

    if not client:
        conn.close()
        return jsonify({'error': 'Client not found'}), 404

    existing = conn.execute(
        'SELECT id, diet, estimated_calories, estimated_protein, estimated_sodium, estimated_saturated_fat, notes '
        'FROM nutrition_logs WHERE client_id = ? AND date = ?',
        (client_id, date)
    ).fetchone()

    if existing and not override:
        conn.close()
        return jsonify({
            'conflict': True,
            'existing': {
                'id': existing['id'],
                'diet': existing['diet'],
                'estimated_calories': existing['estimated_calories'],
                'estimated_protein': existing['estimated_protein'],
                'estimated_sodium': existing['estimated_sodium'],
                'estimated_saturated_fat': existing['estimated_saturated_fat'],
                'notes': existing['notes']
            }
        }), 409

    try:
        if existing and override:
            conn.execute('''
                UPDATE nutrition_logs
                SET diet = ?, estimated_calories = ?, estimated_protein = ?, estimated_sodium = ?,
                    estimated_saturated_fat = ?, notes = ?, updated_at = ?
                WHERE id = ? AND client_id = ?
            ''', (diet, estimated_calories, estimated_protein, estimated_sodium,
                  estimated_saturated_fat, notes, datetime.now(), existing['id'], client_id))
            nutrition_id = existing['id']
        else:
            nutrition_id = str(uuid.uuid4())
            conn.execute('''
                INSERT INTO nutrition_logs
                (id, client_id, date, diet, estimated_calories, estimated_protein, estimated_sodium, estimated_saturated_fat, notes, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (nutrition_id, client_id, date, diet, estimated_calories, estimated_protein, estimated_sodium,
                  estimated_saturated_fat, notes, datetime.now(), datetime.now()))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'id': nutrition_id})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/nutrition-log/<nutrition_id>', methods=['PUT'])
@login_required
def update_nutrition_log(nutrition_id):
    data = request.json
    date = data['date']
    diet = data['diet']
    estimated_calories = data.get('estimated_calories')
    estimated_protein = data.get('estimated_protein')
    estimated_sodium = data.get('estimated_sodium')
    estimated_saturated_fat = data.get('estimated_saturated_fat')
    notes = data.get('notes', '')

    conn = get_db()

    # Verify nutrition log belongs to current trainer's client
    nutrition_log = conn.execute('''
        SELECT nl.* FROM nutrition_logs nl
        JOIN clients c ON nl.client_id = c.id
        WHERE nl.id = ? AND c.trainer_id = ?
    ''', (nutrition_id, session['user_id'])).fetchone()

    if not nutrition_log:
        conn.close()
        return jsonify({'error': 'Nutrition log not found'}), 404

    try:
        conn.execute('''
            UPDATE nutrition_logs
            SET date = ?, diet = ?, estimated_calories = ?, estimated_protein = ?, estimated_sodium = ?,
                estimated_saturated_fat = ?, notes = ?, updated_at = ?
            WHERE id = ?
        ''', (date, diet, estimated_calories, estimated_protein, estimated_sodium, estimated_saturated_fat,
              notes, datetime.now(), nutrition_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/nutrition-log/<nutrition_id>', methods=['DELETE'])
@login_required
def delete_nutrition_log(nutrition_id):
    conn = get_db()

    # Verify nutrition log belongs to current trainer's client
    nutrition_log = conn.execute('''
        SELECT nl.* FROM nutrition_logs nl
        JOIN clients c ON nl.client_id = c.id
        WHERE nl.id = ? AND c.trainer_id = ?
    ''', (nutrition_id, session['user_id'])).fetchone()

    if not nutrition_log:
        conn.close()
        return jsonify({'error': 'Nutrition log not found'}), 404

    try:
        conn.execute('DELETE FROM nutrition_logs WHERE id = ?', (nutrition_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/clients/<client_id>/sleep-logs')
@login_required
def client_sleep_logs(client_id):
    conn = get_db()
    client = conn.execute('SELECT * FROM clients WHERE id = ? AND trainer_id = ?',
                          (client_id, session['user_id'])).fetchone()

    if not client:
        flash('Client not found')
        return redirect(url_for('clients'))

    sleep_history_rows = conn.execute('''
        SELECT id, date, hours, notes
        FROM sleep_logs
        WHERE client_id = ?
        ORDER BY date DESC
    ''', (client_id,)).fetchall()

    app.logger.info(f"[v0] Fetching sleep logs for client {client_id}")
    app.logger.info(f"[v0] Found {len(sleep_history_rows)} sleep log entries")

    sleep_history = [dict(row) for row in sleep_history_rows]

    for entry in sleep_history:
        app.logger.info(f"[v0] Sleep entry: id={entry.get('id')}, date={entry.get('date')}, hours={entry.get('hours')}")

    conn.close()
    return render_template('dashboard/clients/sleep_logs.html',
                           client=dict(client),
                           sleep_history=sleep_history)


@app.route('/api/sleep-log', methods=['POST'])
@login_required
def add_sleep_log():
    data = request.json
    client_id = data['client_id']
    date = data['date']
    hours = data['hours']
    notes = data.get('notes', '')
    override = data.get('override', False)

    conn = get_db()
    client = conn.execute('SELECT * FROM clients WHERE id = ? AND trainer_id = ?',
                          (client_id, session['user_id'])).fetchone()

    if not client:
        conn.close()
        return jsonify({'error': 'Client not found'}), 404

    existing = conn.execute(
        'SELECT id, hours, notes FROM sleep_logs WHERE client_id = ? AND date = ?',
        (client_id, date)
    ).fetchone()

    if existing and not override:
        conn.close()
        return jsonify({
            'conflict': True,
            'existing': {'id': existing['id'], 'hours': existing['hours'], 'notes': existing['notes']}
        }), 409

    try:
        if existing and override:
            conn.execute(
                'UPDATE sleep_logs SET hours = ?, notes = ?, updated_at = ? WHERE id = ? AND client_id = ?',
                (hours, notes, datetime.now(), existing['id'], client_id)
            )
            sleep_id = existing['id']
        else:
            sleep_id = str(uuid.uuid4())
            conn.execute('''
                INSERT INTO sleep_logs (id, client_id, date, hours, notes, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (sleep_id, client_id, date, hours, notes, datetime.now(), datetime.now()))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'id': sleep_id})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/sleep-log/<sleep_id>', methods=['PUT'])
@login_required
def update_sleep_log(sleep_id):
    data = request.json
    date = data['date']
    hours = data['hours']
    notes = data.get('notes', '')

    conn = get_db()

    sleep_log = conn.execute('''
        SELECT sl.* FROM sleep_logs sl
        JOIN clients c ON sl.client_id = c.id
        WHERE sl.id = ? AND c.trainer_id = ?
    ''', (sleep_id, session['user_id'])).fetchone()

    if not sleep_log:
        conn.close()
        return jsonify({'error': 'Sleep log not found'}), 404

    try:
        conn.execute('''
            UPDATE sleep_logs
            SET date = ?, hours = ?, notes = ?, updated_at = ?
            WHERE id = ?
        ''', (date, hours, notes, datetime.now(), sleep_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/sleep-log/<sleep_id>', methods=['DELETE'])
@login_required
def delete_sleep_log(sleep_id):
    conn = get_db()

    sleep_log = conn.execute('''
        SELECT sl.* FROM sleep_logs sl
        JOIN clients c ON sl.client_id = c.id
        WHERE sl.id = ? AND c.trainer_id = ?
    ''', (sleep_id, session['user_id'])).fetchone()

    if not sleep_log:
        conn.close()
        return jsonify({'error': 'Sleep log not found'}), 404

    try:
        conn.execute('DELETE FROM sleep_logs WHERE id = ?', (sleep_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/sleep-logs/<client_id>/clear-all', methods=['DELETE'])
@login_required
def clear_all_sleep_logs(client_id):
    conn = get_db()

    # Verify client belongs to trainer
    client = conn.execute('SELECT * FROM clients WHERE id = ? AND trainer_id = ?',
                          (client_id, session['user_id'])).fetchone()

    if not client:
        conn.close()
        return jsonify({'error': 'Client not found'}), 404

    try:
        result = conn.execute('DELETE FROM sleep_logs WHERE client_id = ?', (client_id,))
        deleted_count = result.rowcount
        conn.commit()
        conn.close()
        app.logger.info(f"[v0] Cleared {deleted_count} sleep logs for client {client_id}")
        return jsonify({'success': True, 'deleted_count': deleted_count})
    except Exception as e:
        conn.close()
        app.logger.error(f"[v0] Error clearing sleep logs: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/sleep-log/import', methods=['POST'])
@login_required
def import_sleep_logs():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    client_id = request.form.get('client_id')

    app.logger.info(f"[v0] Starting sleep log import for client {client_id}")

    if not file or file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    conn = get_db()

    # Verify client belongs to trainer
    client = conn.execute('SELECT * FROM clients WHERE id = ? AND trainer_id = ?',
                          (client_id, session['user_id'])).fetchone()

    if not client:
        conn.close()
        return jsonify({'error': 'Client not found'}), 404

    try:
        filename = secure_filename(file.filename)
        file_ext = os.path.splitext(filename)[1].lower()

        app.logger.info(f"[v0] Processing file: {filename}, extension: {file_ext}")

        entries = []

        if file_ext == '.csv':
            # Parse CSV file
            file_content = file.read().decode('utf-8')
            csv_reader = csv.reader(io.StringIO(file_content))

            for row in csv_reader:
                if len(row) >= 2:
                    date_str = row[0].strip()
                    hours_str = row[1].strip()
                    notes = row[2].strip() if len(row) >= 3 else None

                    # Parse date from MM/DD/YYYY format
                    try:
                        date_obj = datetime.strptime(date_str, '%m/%d/%Y')
                        date_formatted = date_obj.strftime('%Y-%m-%d')
                        hours = float(hours_str)
                        entries.append((date_formatted, hours, notes))
                    except (ValueError, TypeError):
                        continue  # Skip invalid rows

        elif file_ext in ['.xlsx', '.xls']:
            # Parse Excel file
            workbook = load_workbook(file, read_only=True)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=1, values_only=True):
                if row and len(row) >= 2 and row[0] and row[1]:
                    date_val = row[0]
                    hours_val = row[1]
                    notes_val = row[2] if len(row) >= 3 else None

                    # Handle date - could be string or datetime object
                    try:
                        if isinstance(date_val, datetime):
                            date_formatted = date_val.strftime('%Y-%m-%d')
                        elif isinstance(date_val, str):
                            date_obj = datetime.strptime(date_val.strip(), '%m/%d/%Y')
                            date_formatted = date_obj.strftime('%Y-%m-%d')
                        else:
                            continue

                        hours = float(hours_val)
                        notes = str(notes_val).strip() if notes_val else None
                        entries.append((date_formatted, hours, notes))
                    except (ValueError, TypeError):
                        continue  # Skip invalid rows

        else:
            conn.close()
            return jsonify({'error': 'Unsupported file format. Please use CSV or XLSX'}), 400

        if not entries:
            conn.close()
            return jsonify({'error': 'No valid entries found in file'}), 400

        app.logger.info(f"[v0] Parsed {len(entries)} valid entries from file")

        imported_ids = []
        for date, hours, notes in entries:
            # Check if entry already exists for this date
            existing = conn.execute('''
                SELECT id FROM sleep_logs
                WHERE client_id = ? AND date = ?
            ''', (client_id, date)).fetchone()

            if existing:
                conn.execute('''
                    UPDATE sleep_logs
                    SET hours = ?, notes = ?, updated_at = ?
                    WHERE id = ?
                ''', (hours, notes, datetime.now(), existing['id']))
                imported_ids.append(existing['id'])
                app.logger.info(f"[v0] Updated existing entry: id={existing['id']}, date={date}")
            else:
                new_id = str(uuid.uuid4())
                conn.execute('''
                    INSERT INTO sleep_logs (id, client_id, date, hours, notes)
                    VALUES (?, ?, ?, ?, ?)
                ''', (new_id, client_id, date, hours, notes))
                imported_ids.append(new_id)
                app.logger.info(f"[v0] Inserted new entry: id={new_id}, date={date}, hours={hours}")

        conn.commit()
        app.logger.info(f"[v0] Import completed. Total entries processed: {len(imported_ids)}")
        app.logger.info(f"[v0] Imported IDs: {imported_ids}")
        conn.close()

        return jsonify({
            'success': True,
            'message': f'Successfully imported {len(imported_ids)} sleep entries'
        })

    except Exception as e:
        app.logger.error(f"[v0] Import error: {str(e)}")
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500


# ════════════════════════════════════════════════════════
# CLIENT PORTAL — routes live in clients.py
# ════════════════════════════════════════════════════════
from clients import (
    register_client_routes,
    init_client_accounts_table,
    backfill_client_access_codes,
    init_activity_log_table,
    init_nutrition_protein_column,
)

register_client_routes(app)

# Ensure the activity_log table exists even under WSGI (PythonAnywhere never runs
# the __main__ block). CREATE TABLE IF NOT EXISTS is idempotent, so this is safe.
init_activity_log_table()
# Ensure nutrition_logs has the estimated_protein column (idempotent migration).
# This also runs at module level so it applies under WSGI and protects against
# any database file being swapped/imported without the column.
init_nutrition_protein_column()
# Ensure workout_templates has the client_id column (universal vs client-specific).
init_template_client_column()
# Ensure workout_logs has the workout_type column (weightlifting vs cardio).
init_workout_type_column()
# Ensure workout_templates has the workout_type column too (weightlifting vs cardio templates).
init_template_type_column()


if __name__ == '__main__':
    from init_db import init_database
    init_database()
    init_client_accounts_table()
    backfill_client_access_codes()
    init_activity_log_table()
    init_nutrition_protein_column()
    init_template_client_column()
    init_workout_type_column()
    init_template_type_column()
    app.run(debug=True)